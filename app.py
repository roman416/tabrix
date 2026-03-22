from __future__ import annotations

import csv
import json
import logging
import os
import re
import sqlite3
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests
from flask import Flask, jsonify, render_template, request, send_from_directory
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)
CSV_CACHE_DIR = UPLOAD_DIR / '_csv_cache'
CSV_CACHE_DIR.mkdir(exist_ok=True)
DB_PATH = BASE_DIR / 'tabrix.db'

ALLOWED_EXTENSIONS = {'.csv', '.tsv', '.xlsx', '.xls', '.xlsm', '.parquet'}
DEFAULT_OLLAMA_URL = os.getenv('OLLAMA_URL', 'http://127.0.0.1:11434')
DEFAULT_OLLAMA_MODEL = os.getenv('OLLAMA_MODEL', 'deepseek-r1:8b')
MAX_PREVIEW_ROWS = 300
DEFAULT_PREVIEW_ROWS = 50
SORT_CACHE_VERSION = 1

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
logging.basicConfig(level=os.getenv('LOG_LEVEL', 'INFO'))


@dataclass
class TableInfo:
    table_id: str
    name: str
    path: Path
    csv_cache_path: Path
    rows: int
    columns: int
    column_names: list[str]
    dtypes: dict[str, str]
    file_size: int


TABLE_REGISTRY: dict[str, TableInfo] = {}
DATAFRAME_CACHE: dict[str, pd.DataFrame] = {}


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with get_db_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sort_cache (
                table_id TEXT NOT NULL,
                column_name TEXT NOT NULL,
                ascending INTEGER NOT NULL,
                cache_version INTEGER NOT NULL DEFAULT 1,
                row_order_json TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (table_id, column_name, ascending)
            )
        """)
        conn.commit()




init_db()


def invalidate_sort_cache(table_id: str) -> None:
    with get_db_connection() as conn:
        conn.execute('DELETE FROM sort_cache WHERE table_id = ?', (table_id,))
        conn.commit()


def get_cached_sort_order(table_id: str, column: str, ascending: bool, expected_rows: int) -> list[int] | None:
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT row_order_json, row_count, cache_version
            FROM sort_cache
            WHERE table_id = ? AND column_name = ? AND ascending = ?
            """,
            (table_id, column, 1 if ascending else 0),
        ).fetchone()
    if not row:
        return None
    if int(row['row_count']) != int(expected_rows) or int(row['cache_version']) != SORT_CACHE_VERSION:
        invalidate_sort_cache(table_id)
        return None
    try:
        order = json.loads(row['row_order_json'])
    except json.JSONDecodeError:
        invalidate_sort_cache(table_id)
        return None
    if not isinstance(order, list) or len(order) != expected_rows:
        invalidate_sort_cache(table_id)
        return None
    return [int(item) for item in order]


def save_sort_order(table_id: str, column: str, ascending: bool, row_order: list[int]) -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO sort_cache (table_id, column_name, ascending, cache_version, row_order_json, row_count)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(table_id, column_name, ascending) DO UPDATE SET
                cache_version = excluded.cache_version,
                row_order_json = excluded.row_order_json,
                row_count = excluded.row_count,
                created_at = CURRENT_TIMESTAMP
            """,
            (table_id, column, 1 if ascending else 0, SORT_CACHE_VERSION, json.dumps(row_order), len(row_order)),
        )
        conn.commit()


def get_sorted_df(table_id: str, column: str, ascending: bool) -> tuple[pd.DataFrame, str]:
    df = get_df(table_id)
    order = get_cached_sort_order(table_id, column, ascending, len(df))
    if order is not None:
        return df.iloc[order].copy(), 'db_cache'

    sorted_df = df.sort_values(by=column, ascending=ascending, kind='mergesort', na_position='last').copy()
    row_order = [int(idx) for idx in sorted_df.index.tolist()]
    save_sort_order(table_id, column, ascending, row_order)
    return sorted_df, 'computed_and_cached'


PIVOT_AGG_MAP = {
    'count': 'count',
    'unique_count': pd.Series.nunique,
    'sum': 'sum',
    'int_sum': 'sum',
    'mean': 'mean',
    'median': 'median',
    'variance_sample': lambda x: x.var(ddof=1),
    'std_sample': lambda x: x.std(ddof=1),
    'min': 'min',
    'max': 'max',
    'first': 'first',
    'last': 'last',
}

AGGREGATIONS = {
    'count': lambda df, column: df[column].count(),
    'unique_count': lambda df, column: df[column].nunique(dropna=True),
    'sum': lambda df, column: df[column].sum(),
    'int_sum': lambda df, column: int(np.floor(df[column].sum())),
    'mean': lambda df, column: df[column].mean(),
    'median': lambda df, column: df[column].median(),
    'variance_sample': lambda df, column: df[column].var(ddof=1),
    'std_sample': lambda df, column: df[column].std(ddof=1),
    'min': lambda df, column: df[column].min(),
    'max': lambda df, column: df[column].max(),
    'first': lambda df, column: df[column].iloc[0] if not df.empty else None,
    'last': lambda df, column: df[column].iloc[-1] if not df.empty else None,
}


def safe_filename(filename: str) -> str:
    cleaned = re.sub(r'[^\w\-. а-яА-ЯёЁ()]', '_', Path(filename).name).strip('._ ')
    return cleaned or f'table_{uuid.uuid4().hex[:8]}.csv'



def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS



def normalize_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    return value


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on', 'да'}


def summarize_ollama_error(exc: Exception) -> str:
    text = str(exc).strip() or exc.__class__.__name__
    lowered = text.lower()
    if 'failed to establish a new connection' in lowered or 'connection refused' in lowered:
        return 'Ollama не запущена или недоступна по адресу ' + DEFAULT_OLLAMA_URL
    if 'read timed out' in lowered or 'timed out' in lowered:
        return 'Ollama не ответила вовремя. Модель слишком долго думает или зависла.'
    if '404' in lowered and 'generate' in lowered:
        return 'Ollama доступна, но API /api/generate недоступен. Проверь версию Ollama.'
    return text


def get_ollama_status() -> dict[str, Any]:
    try:
        response = requests.get(f'{DEFAULT_OLLAMA_URL}/api/tags', timeout=10)
        response.raise_for_status()
        models = response.json().get('models', [])
        names = [m.get('name') for m in models if m.get('name')]
        return {
            'ok': True,
            'url': DEFAULT_OLLAMA_URL,
            'model': DEFAULT_OLLAMA_MODEL,
            'available_models': names,
            'model_present': DEFAULT_OLLAMA_MODEL in names,
            'error': None,
        }
    except Exception as exc:
        return {
            'ok': False,
            'url': DEFAULT_OLLAMA_URL,
            'model': DEFAULT_OLLAMA_MODEL,
            'error': summarize_ollama_error(exc),
            'model_present': False,
            'available_models': [],
        }



def dataframe_to_records(
    df: pd.DataFrame,
    limit: int = DEFAULT_PREVIEW_ROWS,
    include_row_id: bool = False,
    offset: int = 0,
) -> list[dict[str, Any]]:
    work = df.iloc[offset: offset + limit].copy()
    if include_row_id:
        work.insert(0, '__rowid__', work.index)
    return [{str(k): normalize_value(v) for k, v in row.items()} for row in work.to_dict(orient='records')]



def normalize_header_row(values: tuple[Any, ...] | list[Any] | None) -> list[str]:
    if not values:
        return []
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ''
        if not base:
            base = f'column_{index}'
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f'{base}_{count + 1}')
    return headers



def read_excel_like(file_path: Path) -> pd.DataFrame:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = normalize_header_row(header_row)
        if not headers:
            return pd.DataFrame()

        data_rows: list[list[Any]] = []
        width = len(headers)
        for row in rows_iter:
            values = list(row[:width])
            if len(values) < width:
                values.extend([None] * (width - len(values)))
            data_rows.append(values)
        return pd.DataFrame(data_rows, columns=headers)
    finally:
        workbook.close()



def read_table(file_path: Path) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    if suffix == '.csv':
        try:
            return pd.read_csv(file_path, sep=None, engine='python', encoding='utf-8-sig')
        except Exception:
            return pd.read_csv(file_path, encoding='utf-8-sig')
    if suffix == '.tsv':
        return pd.read_csv(file_path, sep='\t', encoding='utf-8-sig')
    if suffix in {'.xlsx', '.xlsm'}:
        return read_excel_like(file_path)
    if suffix == '.xls':
        return pd.read_excel(file_path, engine='xlrd')
    if suffix == '.parquet':
        return pd.read_parquet(file_path)
    raise ValueError(f'Неподдерживаемый формат: {suffix}')



def value_for_csv_tool(value: Any) -> str:
    if value is None or pd.isna(value):
        return ''
    if isinstance(value, pd.Timestamp):
        text = value.isoformat(sep=' ')
    else:
        text = str(value)
    return text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')



def write_csv_cache(df: pd.DataFrame, csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open('w', newline='', encoding='utf-8-sig') as handle:
        writer = csv.writer(handle)
        writer.writerow([str(col) for col in df.columns])
        for row in df.itertuples(index=False, name=None):
            writer.writerow([value_for_csv_tool(value) for value in row])



def save_table(info: TableInfo, df: pd.DataFrame) -> None:
    suffix = info.path.suffix.lower()
    if suffix == '.csv':
        df.to_csv(info.path, index=False)
    elif suffix == '.tsv':
        df.to_csv(info.path, index=False, sep='\t')
    elif suffix in {'.xlsx', '.xlsm'}:
        df.to_excel(info.path, index=False)
    elif suffix == '.xls':
        fallback_path = info.path.with_suffix('.xlsx')
        df.to_excel(fallback_path, index=False)
        info.path = fallback_path
        info.name = fallback_path.name
    elif suffix == '.parquet':
        df.to_parquet(info.path, index=False)
    else:
        raise ValueError(f'Невозможно сохранить формат: {suffix}')

    write_csv_cache(df, info.csv_cache_path)

    info.rows = int(df.shape[0])
    info.columns = int(df.shape[1])
    info.column_names = [str(c) for c in df.columns]
    info.dtypes = {str(c): str(df[c].dtype) for c in df.columns}
    info.file_size = info.path.stat().st_size if info.path.exists() else info.file_size
    invalidate_sort_cache(info.table_id)



def make_table_info(file_path: Path, df: pd.DataFrame, table_id: str, csv_cache_path: Path) -> TableInfo:
    return TableInfo(
        table_id=table_id,
        name=file_path.name,
        path=file_path,
        csv_cache_path=csv_cache_path,
        rows=int(df.shape[0]),
        columns=int(df.shape[1]),
        column_names=[str(c) for c in df.columns],
        dtypes={str(c): str(df[c].dtype) for c in df.columns},
        file_size=file_path.stat().st_size,
    )



def register_table(file_path: Path) -> TableInfo:
    df = read_table(file_path)
    table_id = uuid.uuid4().hex[:12]
    csv_cache_path = CSV_CACHE_DIR / f'{table_id}.csv'
    write_csv_cache(df, csv_cache_path)
    info = make_table_info(file_path, df, table_id=table_id, csv_cache_path=csv_cache_path)
    TABLE_REGISTRY[info.table_id] = info
    DATAFRAME_CACHE[info.table_id] = df
    return info



def get_df(table_id: str) -> pd.DataFrame:
    if table_id in DATAFRAME_CACHE:
        return DATAFRAME_CACHE[table_id]
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        raise KeyError('Таблица не найдена')
    df = read_table(info.path)
    DATAFRAME_CACHE[table_id] = df
    if not info.csv_cache_path.exists():
        write_csv_cache(df, info.csv_cache_path)
    return df



def info_to_dict(info: TableInfo) -> dict[str, Any]:
    return {
        'table_id': info.table_id,
        'name': info.name,
        'rows': info.rows,
        'columns': info.columns,
        'column_names': info.column_names,
        'dtypes': info.dtypes,
        'file_size': info.file_size,
        'csv_cache': info.csv_cache_path.name,
    }



def detect_column_groups(df: pd.DataFrame) -> dict[str, list[str]]:
    numeric = [str(c) for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    datetime_cols: list[str] = []
    categorical: list[str] = []
    for c in df.columns:
        if str(c) in numeric:
            continue
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            datetime_cols.append(str(c))
        else:
            categorical.append(str(c))
    return {'numeric': numeric, 'categorical': categorical, 'datetime': datetime_cols}



def series_stats(series: pd.Series) -> dict[str, Any]:
    s = series.dropna()
    if s.empty:
        return {
            'count': 0,
            'unique_count': 0,
            'unique_values': [],
            'sum': None,
            'int_sum': None,
            'mean': None,
            'median': None,
            'variance_sample': None,
            'std_sample': None,
            'min': None,
            'max': None,
            'first': None,
            'last': None,
        }

    result = {
        'count': int(s.shape[0]),
        'unique_count': int(s.nunique(dropna=True)),
        'unique_values': s.astype(str).drop_duplicates().head(50).tolist(),
        'sum': None,
        'int_sum': None,
        'mean': None,
        'median': None,
        'variance_sample': None,
        'std_sample': None,
        'min': normalize_value(s.min()),
        'max': normalize_value(s.max()),
        'first': normalize_value(s.iloc[0]),
        'last': normalize_value(s.iloc[-1]),
    }
    if pd.api.types.is_numeric_dtype(s):
        result.update(
            {
                'sum': normalize_value(s.sum()),
                'int_sum': int(np.floor(s.sum())),
                'mean': normalize_value(s.mean()),
                'median': normalize_value(s.median()),
                'variance_sample': normalize_value(s.var(ddof=1)) if len(s) > 1 else None,
                'std_sample': normalize_value(s.std(ddof=1)) if len(s) > 1 else None,
                'min': normalize_value(s.min()),
                'max': normalize_value(s.max()),
            }
        )
    return result



def coerce_value_for_series(raw_value: Any, series: pd.Series) -> Any:
    if raw_value is None:
        return None
    if isinstance(raw_value, str):
        stripped = raw_value.strip()
        if stripped == '':
            return None
        raw_value = stripped

    if pd.api.types.is_integer_dtype(series.dtype):
        return int(float(raw_value))
    if pd.api.types.is_float_dtype(series.dtype):
        return float(raw_value)
    if pd.api.types.is_bool_dtype(series.dtype):
        if isinstance(raw_value, bool):
            return raw_value
        return str(raw_value).strip().lower() in {'1', 'true', 'yes', 'да'}
    if pd.api.types.is_datetime64_any_dtype(series.dtype):
        return pd.to_datetime(raw_value)
    return raw_value



def restore_typed_value(raw_value: str | None, series: pd.Series) -> Any:
    if raw_value is None or raw_value == '':
        return None
    if pd.api.types.is_integer_dtype(series.dtype):
        return int(float(raw_value))
    if pd.api.types.is_float_dtype(series.dtype):
        return float(raw_value)
    if pd.api.types.is_bool_dtype(series.dtype):
        return raw_value.strip().lower() in {'1', 'true', 'yes', 'да'}
    if pd.api.types.is_datetime64_any_dtype(series.dtype):
        return pd.to_datetime(raw_value).isoformat()
    return raw_value



def strip_think_blocks(text: str) -> str:
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL | re.IGNORECASE)
    return text.strip()





def column_label_to_index(label: str) -> int:
    label = re.sub(r'[^A-Za-zА-Яа-я]', '', str(label or '')).upper()
    if not label:
        raise ValueError('Пустая координата колонки')
    # Поддерживаем латиницу для координат A, B, AA...
    if not re.fullmatch(r'[A-Z]+', label):
        raise ValueError('Координата колонки должна быть в формате A, B, AA')
    index = 0
    for ch in label:
        index = index * 26 + (ord(ch) - ord('A') + 1)
    return index - 1


def parse_cell_reference(cell_ref: str, df: pd.DataFrame) -> tuple[int, str, int]:
    match = re.fullmatch(r'\s*([A-Za-z]+)\s*(\d+)\s*', str(cell_ref or ''))
    if not match:
        raise ValueError('Координата должна быть в формате A1')
    column_label, row_label = match.groups()
    col_index = column_label_to_index(column_label)
    if col_index < 0 or col_index >= len(df.columns):
        raise ValueError('Колонка вне диапазона таблицы')
    row_id = int(row_label) - 1
    if row_id < 0 or row_id >= len(df.index):
        raise ValueError('Строка вне диапазона таблицы')
    return row_id, str(df.columns[col_index]), col_index


def safe_compare_as_number(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def apply_simple_filter(df: pd.DataFrame, expression: str) -> tuple[pd.DataFrame, str]:
    expr = str(expression or '').strip()
    pattern = re.compile(r'^\s*(?:ячейка|столбец|колонка)?\s*([A-Za-zА-Яа-я][\wа-яА-Я]*)\s*(>=|<=|!=|=|>|<|contains|содержит)\s*(.+?)\s*$', re.IGNORECASE)
    match = pattern.match(expr)
    if not match:
        raise ValueError('Не удалось разобрать фильтр. Пример: A>100 или amount contains abc')
    raw_column, operator, raw_value = match.groups()
    operator = operator.lower()
    raw_value = raw_value.strip().strip("'\"")

    if re.fullmatch(r'[A-Za-z]+', raw_column):
        col_index = column_label_to_index(raw_column)
        if col_index < 0 or col_index >= len(df.columns):
            raise ValueError('Колонка вне диапазона таблицы')
        column = str(df.columns[col_index])
    else:
        column = raw_column
        if column not in df.columns:
            raise ValueError('Колонка не найдена')

    series = df[column]
    numeric_value = safe_compare_as_number(raw_value)
    numeric_series = pd.to_numeric(series, errors='coerce')

    if operator in {'contains', 'содержит'}:
        mask = series.astype(str).str.contains(raw_value, case=False, na=False)
    elif operator in {'>', '<', '>=', '<='} and numeric_value is not None:
        ops = {
            '>': numeric_series > numeric_value,
            '<': numeric_series < numeric_value,
            '>=': numeric_series >= numeric_value,
            '<=': numeric_series <= numeric_value,
        }
        mask = ops[operator].fillna(False)
    elif operator in {'=', '!='}:
        if numeric_value is not None:
            mask = (numeric_series == numeric_value) if operator == '=' else (numeric_series != numeric_value)
            mask = mask.fillna(False)
        else:
            compare = series.astype(str).str.lower()
            target = str(raw_value).lower()
            mask = compare.eq(target) if operator == '=' else ~compare.eq(target)
    else:
        raise ValueError('Для такого оператора сейчас поддерживается только сравнение с числом или contains')

    return df.loc[mask].copy(), column


def build_llm_context(table_id: str | None, question: str) -> dict[str, Any]:
    context: dict[str, Any] = {
        'table': None,
        'preview': None,
        'column_groups': None,
        'stats': {},
        'cell_lookup': None,
        'page_lookup': None,
        'row_lookup': None,
        'filter_stats': None,
        'ollama': {'url': DEFAULT_OLLAMA_URL, 'model': DEFAULT_OLLAMA_MODEL},
    }
    if not table_id or table_id not in TABLE_REGISTRY:
        return context

    df = get_df(table_id)
    context['table'] = info_to_dict(TABLE_REGISTRY[table_id])
    context['preview'] = dataframe_to_records(df, 10)
    context['column_groups'] = detect_column_groups(df)
    for col in detect_column_groups(df)['numeric'][:3]:
        context['stats'][col] = series_stats(df[col])

    q = question.strip()

    cell_match = re.search(r'([A-Za-z]+\d+)', q)
    if cell_match:
        try:
            row_id, column, col_index = parse_cell_reference(cell_match.group(1), df)
            context['cell_lookup'] = {
                'cell': cell_match.group(1).upper(),
                'row_id': row_id,
                'column': column,
                'column_index': col_index,
                'value': normalize_value(df.iloc[row_id, col_index]),
                'row': {str(k): normalize_value(v) for k, v in df.iloc[row_id].items()},
            }
        except Exception:
            pass

    page_match = re.search(r'страниц[аеиуы]?\s*(\d+)', q, flags=re.IGNORECASE)
    if page_match:
        page = max(int(page_match.group(1)), 1)
        offset = (page - 1) * 100
        page_df = df.iloc[offset: offset + 100].copy()
        context['page_lookup'] = {
            'page': page,
            'page_size': 100,
            'offset': offset,
            'rows_on_page': int(page_df.shape[0]),
            'preview': dataframe_to_records(page_df, 100, include_row_id=True),
        }

    filter_match = re.search(r'где\s+(.+)', q, flags=re.IGNORECASE)
    if filter_match:
        try:
            filtered, used_column = apply_simple_filter(df, filter_match.group(1))
            numeric_cols = [str(c) for c in filtered.columns if pd.api.types.is_numeric_dtype(filtered[c])]
            context['filter_stats'] = {
                'filter': filter_match.group(1).strip(),
                'column': used_column,
                'rows': int(filtered.shape[0]),
                'preview': dataframe_to_records(filtered, 25, include_row_id=True),
                'numeric_stats': {col: series_stats(filtered[col]) for col in numeric_cols[:5]},
            }
        except Exception as exc:
            context['filter_stats'] = {'filter': filter_match.group(1).strip(), 'error': str(exc)}

    row_search_patterns = [
        re.search(r'(?:строк[ауи]|ряд)\s+где\s+([A-Za-zА-Яа-я][\wа-яА-Я]*)\s*=\s*([^?]+)', q, flags=re.IGNORECASE),
        re.search(r'найд[ии].*?([A-Za-zА-Яа-я][\wа-яА-Я]*)\s*=\s*([^?]+)', q, flags=re.IGNORECASE),
    ]
    for m in row_search_patterns:
        if not m:
            continue
        raw_column, raw_value = m.groups()
        try:
            if re.fullmatch(r'[A-Za-z]+', raw_column):
                col_index = column_label_to_index(raw_column)
                column = str(df.columns[col_index])
            else:
                column = raw_column
            if column in df.columns:
                target = raw_value.strip().strip("'\"")
                mask = df[column].astype(str).str.lower() == target.lower()
                found = df.loc[mask].head(5)
                context['row_lookup'] = {
                    'column': column,
                    'value': target,
                    'matches': int(mask.sum()),
                    'rows': dataframe_to_records(found, 5, include_row_id=True),
                }
                break
        except Exception:
            pass

    return context

def table_overview_payload(table_id: str) -> dict[str, Any]:
    info = TABLE_REGISTRY[table_id]
    df = get_df(table_id)
    return {
        'table': info_to_dict(info),
        'preview': dataframe_to_records(df, DEFAULT_PREVIEW_ROWS, include_row_id=True),
        'column_groups': detect_column_groups(df),
    }


@app.get('/')
def index():
    return render_template(
        'app.html',
        title='Табрикс · Обзор',
        page_name='overview',
        section_kicker='Главная панель',
        section_title='Редактор и обзор таблицы',
        section_description='Загружай файлы, правь ячейки, фильтруй строки и считай статистику без плясок вокруг вкладок.',
        initial_tab='tables',
    )


@app.get('/dashboards')
def dashboards_page():
    return render_template(
        'app.html',
        title='Табрикс · Дашборды',
        page_name='dashboards',
        section_kicker='Визуализация',
        section_title='Автоматические дашборды',
        section_description='Ключевые графики и аналитические карточки по активной таблице на отдельной странице, как и положено взрослому интерфейсу.',
        initial_tab='dashboards',
    )


@app.get('/assistant')
def assistant_page():
    return render_template(
        'app.html',
        title='Табрикс · LLM',
        page_name='assistant',
        section_kicker='Локальная модель',
        section_title='Помощник по данным',
        section_description='Вопросы к таблице через fallback-логику или Ollama, когда она не решила умереть именно сегодня.',
        initial_tab='dashboards',
    )


@app.get('/health')
def health():
    return jsonify({'status': 'ok', 'tables': len(TABLE_REGISTRY), 'ollama_model': DEFAULT_OLLAMA_MODEL})


@app.post('/api/upload')
def upload_table():
    file = request.files.get('file')
    if file is None or not file.filename:
        return jsonify({'error': 'Файл не передан'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Поддерживаются CSV, TSV, XLSX, XLS, XLSM, PARQUET'}), 400

    filename = safe_filename(file.filename)
    target = UPLOAD_DIR / filename
    if target.exists():
        target = UPLOAD_DIR / f'{target.stem}_{uuid.uuid4().hex[:6]}{target.suffix}'
    file.save(target)

    try:
        info = register_table(target)
    except Exception as exc:
        target.unlink(missing_ok=True)
        return jsonify({'error': f'Не удалось прочитать таблицу: {exc}'}), 400

    df = DATAFRAME_CACHE[info.table_id]
    return jsonify({
        'message': 'Таблица загружена.',
        'table': info_to_dict(info),
        'preview': dataframe_to_records(df, 15, include_row_id=True),
        'column_groups': detect_column_groups(df),
    })


@app.get('/api/tables')
def list_tables():
    tables = [info_to_dict(info) for info in TABLE_REGISTRY.values()]
    return jsonify({'tables': tables})


@app.get('/api/table/<table_id>')
def table_details(table_id: str):
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404
    return jsonify(table_overview_payload(table_id))


@app.get('/api/table/<table_id>/preview')
def table_preview(table_id: str):
    df = get_df(table_id)
    limit = min(max(int(request.args.get('limit', DEFAULT_PREVIEW_ROWS)), 1), MAX_PREVIEW_ROWS)
    offset = max(int(request.args.get('offset', 0)), 0)
    query = str(request.args.get('query', '')).strip().lower()

    sort_column = str(request.args.get('sort_column', '')).strip()
    sort_direction = str(request.args.get('sort_direction', 'asc')).strip().lower()
    ascending = sort_direction != 'desc'

    work = df
    sort_source = None
    if sort_column:
        if sort_column not in df.columns:
            return jsonify({'error': 'Колонка сортировки не найдена'}), 404
        work, sort_source = get_sorted_df(table_id, sort_column, ascending)

    if query:
        mask = pd.Series(False, index=work.index)
        for col in work.columns:
            mask = mask | work[col].astype(str).str.lower().str.contains(query, na=False)
        work = work.loc[mask]

    return jsonify({
        'rows': dataframe_to_records(work, limit, include_row_id=True, offset=offset),
        'total_rows': int(work.shape[0]),
        'offset': offset,
        'limit': limit,
        'query': query,
        'sort': {'column': sort_column or None, 'ascending': ascending} if sort_column else None,
        'sort_source': sort_source,
    })


@app.get('/api/table/<table_id>/stats/<column>')
def column_stats(table_id: str, column: str):
    df = get_df(table_id)
    if column not in df.columns:
        return jsonify({'error': 'Колонка не найдена'}), 404

    stats = series_stats(df[column])
    return jsonify({'column': column, 'stats': stats, 'source': 'python'})


@app.post('/api/table/<table_id>/sort')
def sort_table(table_id: str):
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404

    df = get_df(table_id)
    payload = request.get_json(silent=True) or {}
    column = payload.get('column')
    ascending = bool(payload.get('ascending', True))

    if not column or column not in df.columns:
        return jsonify({'error': 'Укажи корректную колонку сортировки'}), 400

    sorted_df, source = get_sorted_df(table_id, column, ascending)
    return jsonify({
        'message': 'Сортировка готова и сохранена в базе данных.',
        'preview': dataframe_to_records(sorted_df, 25, include_row_id=True),
        'source': source,
        'sort': {'column': column, 'ascending': ascending},
        'table': info_to_dict(info),
    })


@app.post('/api/table/<table_id>/pivot')
def build_pivot(table_id: str):
    df = get_df(table_id).copy()
    payload = request.get_json(silent=True) or {}
    rows = [c for c in payload.get('rows', []) if c in df.columns]
    columns = [c for c in payload.get('columns', []) if c in df.columns]
    value = payload.get('value')
    aggfunc = payload.get('aggfunc', 'sum')

    if not rows and not columns:
        return jsonify({'error': 'Нужна хотя бы одна группировка: rows или columns'}), 400
    if value not in df.columns:
        return jsonify({'error': 'Неверная колонка value'}), 400
    if aggfunc not in AGGREGATIONS:
        return jsonify({'error': f'Неизвестная агрегация: {aggfunc}'}), 400

    pt = pd.pivot_table(
        df,
        index=rows or None,
        columns=columns or None,
        values=value,
        aggfunc=PIVOT_AGG_MAP[aggfunc],
        fill_value=0,
        dropna=False,
        margins=True,
        margins_name='Итого',
    )

    if aggfunc == 'int_sum':
        pt = np.floor(pt).astype(int)

    result = pt.reset_index()
    result.columns = [
        ' | '.join(str(part) for part in col if str(part) != '').strip() if isinstance(col, tuple) else str(col)
        for col in result.columns
    ]

    return jsonify({
        'rows': dataframe_to_records(result, 500),
        'columns': [str(c) for c in result.columns],
        'row_fields': rows,
        'column_fields': columns,
        'value_field': value,
        'aggfunc': aggfunc,
    })


@app.post('/api/table/<table_id>/cell')
def update_cell(table_id: str):
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404

    df = get_df(table_id)
    payload = request.get_json(silent=True) or {}
    row_id = payload.get('row_id')
    column = payload.get('column')
    value = payload.get('value')

    if column not in df.columns:
        return jsonify({'error': 'Колонка не найдена'}), 404
    if row_id is None:
        return jsonify({'error': 'Не указан row_id'}), 400

    try:
        row_id = int(row_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'row_id должен быть числом'}), 400

    if row_id not in df.index:
        return jsonify({'error': 'Строка не найдена'}), 404

    try:
        coerced_value = coerce_value_for_series(value, df[column])
        df.at[row_id, column] = coerced_value
        DATAFRAME_CACHE[table_id] = df
        save_table(info, df)
    except Exception as exc:
        return jsonify({'error': f'Не удалось сохранить значение: {exc}'}), 400

    return jsonify({
        'message': 'Ячейка обновлена',
        'row_id': row_id,
        'column': column,
        'value': normalize_value(df.at[row_id, column]),
        'table': info_to_dict(info),
    })


@app.post('/api/table/<table_id>/row')
def add_row(table_id: str):
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404
    df = get_df(table_id)
    new_row = {col: None for col in df.columns}
    df.loc[len(df)] = new_row
    DATAFRAME_CACHE[table_id] = df
    save_table(info, df)
    return jsonify({'message': 'Строка добавлена', 'row_id': int(df.index[-1]), 'table': info_to_dict(info)})


@app.delete('/api/table/<table_id>/row/<int:row_id>')
def delete_row(table_id: str, row_id: int):
    info = TABLE_REGISTRY.get(table_id)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404
    df = get_df(table_id)
    if row_id not in df.index:
        return jsonify({'error': 'Строка не найдена'}), 404
    df = df.drop(index=row_id).reset_index(drop=True)
    DATAFRAME_CACHE[table_id] = df
    save_table(info, df)
    return jsonify({'message': 'Строка удалена', 'row_id': row_id, 'table': info_to_dict(info)})


@app.delete('/api/table/<table_id>')
def delete_table(table_id: str):
    info = TABLE_REGISTRY.pop(table_id, None)
    DATAFRAME_CACHE.pop(table_id, None)
    if not info:
        return jsonify({'error': 'Таблица не найдена'}), 404
    invalidate_sort_cache(table_id)
    info.path.unlink(missing_ok=True)
    info.csv_cache_path.unlink(missing_ok=True)
    return jsonify({'message': 'Таблица удалена', 'table_id': table_id})


@app.get('/api/table/<table_id>/dashboard')
def dashboard(table_id: str):
    df = get_df(table_id)
    groups = detect_column_groups(df)
    numeric = groups['numeric']
    categorical = groups['categorical']

    cards: list[dict[str, Any]] = []
    if numeric:
        first_num = numeric[0]
        cards.append({'title': f'Сводка по {first_num}', 'type': 'stats', 'payload': series_stats(df[first_num])})
    if categorical:
        first_cat = categorical[0]
        top = df[first_cat].astype(str).fillna('(пусто)').value_counts().head(10)
        cards.append({
            'title': f'Топ значений: {first_cat}',
            'type': 'bar',
            'payload': {'labels': top.index.tolist(), 'values': [int(v) for v in top.values.tolist()]},
        })
    if numeric and categorical:
        first_num = numeric[0]
        first_cat = categorical[0]
        agg = df.groupby(first_cat, dropna=False)[first_num].sum().sort_values(ascending=False).head(10)
        cards.append({
            'title': f'Сумма {first_num} по {first_cat}',
            'type': 'bar',
            'payload': {
                'labels': [str(x) for x in agg.index.tolist()],
                'values': [float(v) if pd.notna(v) else 0 for v in agg.values.tolist()],
            },
        })

    return jsonify({'table': info_to_dict(TABLE_REGISTRY[table_id]), 'cards': cards})


@app.post('/api/assistant')
def assistant():
    payload = request.get_json(silent=True) or {}
    table_id = payload.get('table_id')
    question = str(payload.get('question', '')).strip()
    use_ollama = parse_bool(payload.get('use_ollama'), True)

    if not question:
        return jsonify({'error': 'Вопрос пустой'}), 400

    context = build_llm_context(table_id, question)
    fallback = build_fallback_answer(question, context)

    if not use_ollama:
        return jsonify({'answer': fallback, 'source': 'python_fallback', 'model': None, 'warning': 'Используется локальный fallback без модели.'})

    status = get_ollama_status()
    if not status.get('ok'):
        return jsonify({
            'answer': fallback,
            'source': 'python_fallback',
            'model': DEFAULT_OLLAMA_MODEL,
            'warning': status.get('error') or 'Ollama недоступна.',
        })

    if not status.get('model_present'):
        return jsonify({
            'answer': fallback,
            'source': 'python_fallback',
            'model': DEFAULT_OLLAMA_MODEL,
            'warning': f"Модель {DEFAULT_OLLAMA_MODEL} не найдена в Ollama. Доступные модели: {', '.join(status.get('available_models', [])) or 'нет'}",
        })

    try:
        answer = ask_ollama(question, context)
        return jsonify({'answer': answer, 'source': 'ollama', 'model': DEFAULT_OLLAMA_MODEL})
    except Exception as exc:
        logging.exception('Ошибка при запросе к Ollama')
        return jsonify({
            'answer': fallback,
            'source': 'python_fallback',
            'model': DEFAULT_OLLAMA_MODEL,
            'warning': summarize_ollama_error(exc),
        })


@app.get('/api/assistant/status')
def assistant_status():
    return jsonify(get_ollama_status())



def build_fallback_answer(question: str, context: dict[str, Any]) -> str:
    q = question.lower()
    table = context.get('table')
    if not table:
        return 'Сейчас у меня нет выбранной таблицы. Загрузи файл, и я смогу отвечать по колонкам, агрегатам, координатам ячеек и фильтрам.'

    groups = context.get('column_groups') or {}
    numeric = groups.get('numeric', [])
    categorical = groups.get('categorical', [])

    if context.get('cell_lookup') and any(token in q for token in ['ячейк', 'координат', 'значение', 'строк']) :
        cell = context['cell_lookup']
        return (
            f"Ячейка {cell['cell']} это строка {cell['row_id'] + 1}, колонка {cell['column']}. "
            f"Значение: {cell['value']}. Вся строка: {json.dumps(cell['row'], ensure_ascii=False)}"
        )

    if context.get('page_lookup') and 'страниц' in q:
        page = context['page_lookup']
        return (
            f"Страница {page['page']} показывает строки {page['offset'] + 1}-{page['offset'] + page['rows_on_page']} "
            f"при размере страницы 100. Данные страницы: {json.dumps(page['preview'], ensure_ascii=False)}"
        )

    if context.get('row_lookup') and any(token in q for token in ['строк', 'ряд', 'найди']):
        row = context['row_lookup']
        return f"Найдено строк: {row['matches']} по условию {row['column']} = {row['value']}. Первые совпадения: {json.dumps(row['rows'], ensure_ascii=False)}"

    if context.get('filter_stats') and 'где' in q:
        filt = context['filter_stats']
        if filt.get('error'):
            return f"Фильтр распознан, но посчитать не вышло: {filt['error']}"
        return (
            f"По фильтру «{filt['filter']}» найдено строк: {filt['rows']}. "
            f"Предпросмотр: {json.dumps(filt['preview'], ensure_ascii=False)}. "
            f"Статистика по числовым колонкам: {json.dumps(filt['numeric_stats'], ensure_ascii=False)}"
        )

    if 'колон' in q or 'столб' in q:
        return f"В таблице {table['columns']} колонок: {', '.join(table['column_names'][:20])}."
    if 'числов' in q:
        return f"Числовые колонки: {', '.join(numeric) if numeric else 'не найдены'}."
    if 'категор' in q or 'текст' in q:
        return f"Текстовые колонки: {', '.join(categorical) if categorical else 'не найдены'}."
    if 'строк' in q:
        return f"В таблице {table['rows']} строк."
    if numeric:
        first = numeric[0]
        stats = context['stats'].get(first, {})
        return (
            f"Быстрый ответ по таблице {table['name']}: строк {table['rows']}, колонок {table['columns']}. "
            f"Например, по колонке {first}: сумма={stats.get('sum')}, среднее={stats.get('mean')}, минимум={stats.get('min')}, максимум={stats.get('max')}."
        )
    return f"Таблица {table['name']} загружена. В ней {table['rows']} строк и {table['columns']} колонок."


def ask_ollama(question: str, context: dict[str, Any]) -> str:
    prompt = (
        'Ты локальный аналитический помощник табличного приложения. '
        'Отвечай по-русски, кратко и предметно. Не показывай скрытые рассуждения и не пиши служебные теги. '
        'Если данных недостаточно, честно так и скажи.\n\n'
        f"Контекст таблицы:\n{json.dumps(context, ensure_ascii=False, indent=2)}\n\n"
        f'Вопрос пользователя: {question}'
    )

    payload = {
        'model': DEFAULT_OLLAMA_MODEL,
        'prompt': prompt,
        'stream': False,
        'options': {
            'temperature': 0.2,
        },
    }
    response = requests.post(f'{DEFAULT_OLLAMA_URL}/api/generate', json=payload, timeout=180)
    response.raise_for_status()
    data = response.json()

    answer = data.get('response')
    if answer is None and isinstance(data.get('message'), dict):
        answer = data['message'].get('content')
    if answer is None and 'messages' in data and isinstance(data['messages'], list) and data['messages']:
        last_message = data['messages'][-1]
        if isinstance(last_message, dict):
            answer = last_message.get('content')

    cleaned = strip_think_blocks(str(answer or '')).strip()
    if not cleaned:
        raise RuntimeError('Ollama вернула пустой ответ.')
    return cleaned


@app.get('/download/<path:filename>')
def download_file(filename: str):
    return send_from_directory(UPLOAD_DIR, Path(filename).name, as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
